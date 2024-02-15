from openpyxl import load_workbook  # load_workbook permite abrir e carregar um arquivo Excel existente para manipulação.

fileName = "ListaGeral.xlsx"

workbook = load_workbook(fileName)
mainTabSheet = workbook["planilhaGeral"]

colecao = []
count = 0


row_values = mainTabSheet.iter_rows(min_row=0,max_row=3404,min_col=0,max_col=11,values_only=True) # lê a planilha da primeira linha e coluna até a linha 3404 e coluna 11
row_values_list = list(row_values) # converte a tupla em uma lista python
colecao.append(row_values_list)

#print(colecao[0][0][8]) 

for i in range (0,3403):
    if colecao[0][i][8] == "Sim": # o primeiro [0] acessa todos os elementos da coleção pois estes estão dentro do mesmo cochete, enquanto [i] acessa a linha e [9] acessa a coluna
        count += 1



print("count = " + str(count))
print(colecao[0][0][8]) # imprime a linha Aparelho gela?
print(colecao[0][3403][1]) # imprime a ultima linha na coluna com numero de patrimônio
workbook.close()
print("---fim do programa------")

# a planilha possui 3404 linhas